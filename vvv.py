import asyncio
import re

from aiogram import Bot, Dispatcher, types
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart
from aiogram.client.default import DefaultBotProperties
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton

from docx import Document
from openpyxl import load_workbook, Workbook
from PyPDF2 import PdfReader
import xlrd


API_TOKEN = "8697966421:AAGqq4JimRDrjP0rswZCZz92U1gYYQtROao"
ADMIN_ID = 8027087107
CHANNEL = "@krilchadan_lotinchaga"


bot = Bot(token=API_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher()


# ================= USERS =================

def get_users():
    try:
        with open("users.txt","r") as f:
            return set(int(x.strip()) for x in f.readlines())
    except:
        return set()


def save_user(user_id):

    users = get_users()

    if user_id not in users:
        with open("users.txt","a") as f:
            f.write(str(user_id)+"\n")


# ================= KIRIL → LOTIN =================

kiril_map = {
"а":"a","б":"b","в":"v","г":"g","д":"d","е":"e",
"ё":"yo","ж":"j","з":"z","и":"i","й":"y","к":"k",
"қ":"q","л":"l","м":"m","н":"n","о":"o","п":"p",
"р":"r","с":"s","т":"t","у":"u","ф":"f","х":"x",
"ҳ":"h","ч":"ch","ш":"sh","ю":"yu","я":"ya",
"ғ":"g'","ў":"o'","ы":"i","э":"e"
}

def kiril_lotin(text):

    for k,v in kiril_map.items():
        text=text.replace(k,v)
        text=text.replace(k.upper(),v.upper())

    return text


# ================= LOTIN → KIRIL =================

def lotin_kiril(text):

    text=text.lower()

    text=text.replace("’","'")
    text=text.replace("ʻ","'")

    text=re.sub(r'\be','э',text)

    combos={
        "o'":"ў",
        "g'":"ғ",
        "sh":"ш",
        "ch":"ч",
        "ng":"нг",
        "yo":"ё",
        "yu":"ю",
        "ya":"я"
    }

    for k,v in combos.items():
        text=text.replace(k,v)

    letters={
        "a":"а","b":"б","d":"д","e":"е","f":"ф","g":"г",
        "h":"ҳ","i":"и","j":"ж","k":"к","l":"л","m":"м",
        "n":"н","o":"о","p":"п","q":"қ","r":"р","s":"с",
        "t":"т","u":"у","v":"в","x":"х","y":"й","z":"з"
    }

    for k,v in letters.items():
        text=text.replace(k,v)

    return text


def convert(text,mode):

    if mode=="kl":
        return kiril_lotin(text)
    else:
        return lotin_kiril(text)


# ================= USER MENU =================

menu = ReplyKeyboardMarkup(
keyboard=[
[
KeyboardButton(text="🔤 Kiril → Lotin"),
KeyboardButton(text="🔤 Lotin → Kiril")
]
],
resize_keyboard=True
)


# ================= ADMIN MENU =================

admin_menu = InlineKeyboardMarkup(
inline_keyboard=[
[InlineKeyboardButton(text="📊 Statistika",callback_data="admin_stats")],
[InlineKeyboardButton(text="👥 Users",callback_data="admin_users")],
[InlineKeyboardButton(text="📢 Broadcast",callback_data="admin_broadcast")]
]
)


user_mode={}


# ================= START =================

@dp.message(CommandStart())
async def start(message: types.Message):

    user_id=message.from_user.id

    save_user(user_id)

    member=await bot.get_chat_member(CHANNEL,user_id)

    if member.status not in ["member","administrator","creator"]:

        btn=InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="📢 Kanalga o'tish",
                url=f"https://t.me/{CHANNEL.replace('@','')}")]
            ]
        )

        await message.answer(
            "Botdan foydalanish uchun kanalga obuna bo‘ling",
            reply_markup=btn
        )

        return

    await message.answer(
        "👋 <b>Assalomu alaykum!</b>\n\n"
        "🤖 Bu bot yordamida siz:\n\n"
        "🔤 <b>Kiril → Lotin</b>\n"
        "🔤 <b>Lotin → Kiril</b>\n\n"
        "transliteratsiya qilishingiz mumkin.\n\n"
        "📂 Qo‘llab-quvvatlanadigan fayllar:\n"
        "• TXT\n"
        "• Word (.docx)\n"
        "• Excel (.xlsx .xls)\n"
        "• PDF\n\n"
        "✍️ Matn yoki fayl yuboring.",
        reply_markup=menu
    )


# ================= MODE =================

@dp.message(lambda m: m.text=="🔤 Kiril → Lotin")
async def kl_mode(message: types.Message):

    user_mode[message.from_user.id]="kl"

    await message.answer("Kiril matn yuboring")


@dp.message(lambda m: m.text=="🔤 Lotin → Kiril")
async def lk_mode(message: types.Message):

    user_mode[message.from_user.id]="lk"

    await message.answer("Lotin matn yuboring")


# ================= ADMIN PANEL =================

@dp.message(lambda m: m.text=="/admin")
async def admin_panel(message: types.Message):

    if message.from_user.id!=ADMIN_ID:
        return

    await message.answer(
        "⚙️ Admin panel",
        reply_markup=admin_menu
    )


@dp.callback_query()
async def admin_callbacks(call: types.CallbackQuery):

    if call.from_user.id!=ADMIN_ID:
        return


    if call.data=="admin_stats":

        users=get_users()

        await call.message.answer(
            f"👥 Users: {len(users)}"
        )


    elif call.data=="admin_users":

        users=get_users()

        text="\n".join(str(u) for u in users)

        with open("users_list.txt","w") as f:
            f.write(text)

        await call.message.answer_document(types.FSInputFile("users_list.txt"))


    elif call.data=="admin_broadcast":

        await call.message.answer(
            "Yuborish uchun:\n/send matn"
        )


# ================= BROADCAST =================

@dp.message(lambda m: m.text and m.text.startswith("/send"))
async def broadcast(message: types.Message):

    if message.from_user.id!=ADMIN_ID:
        return

    text=message.text.replace("/send ","")

    users=get_users()

    for user in users:

        try:
            await bot.send_message(user,text)
        except:
            pass

    await message.answer("Xabar yuborildi")


# ================= FILE =================

@dp.message(lambda msg: msg.document)
async def file_handler(message: types.Message):

    user_id=message.from_user.id

    mode=user_mode.get(user_id,"kl")

    file_id=message.document.file_id
    file_name=message.document.file_name

    await message.answer("Fayl qabul qilindi...")

    file=await bot.get_file(file_id)

    await bot.download_file(file.file_path,file_name)

    ext=file_name.split(".")[-1].lower()


    if ext=="txt":

        text=open(file_name,"r",encoding="utf-8",errors="ignore").read()

        result=convert(text,mode)

        open("result.txt","w",encoding="utf-8").write(result)

        await message.answer_document(types.FSInputFile("result.txt"))


    elif ext=="docx":

        doc=Document(file_name)

        for p in doc.paragraphs:
            p.text=convert(p.text,mode)

        doc.save("result.docx")

        await message.answer_document(types.FSInputFile("result.docx"))


    elif ext in ["xlsx","xlsm"]:

        wb=load_workbook(file_name)

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value,str):
                        cell.value=convert(cell.value,mode)

        wb.save("result.xlsx")

        await message.answer_document(types.FSInputFile("result.xlsx"))


    elif ext=="xls":

        book=xlrd.open_workbook(file_name)

        new_wb=Workbook()

        sheet=new_wb.active

        r_i=1

        for sh in book.sheets():

            for r in range(sh.nrows):

                c_i=1

                for c in range(sh.ncols):

                    val=sh.cell_value(r,c)

                    if isinstance(val,str):
                        val=convert(val,mode)

                    sheet.cell(row=r_i,column=c_i,value=val)

                    c_i+=1

                r_i+=1

        new_wb.save("result.xlsx")

        await message.answer_document(types.FSInputFile("result.xlsx"))


    elif ext=="pdf":

        reader=PdfReader(file_name)

        text=""

        for page in reader.pages:
            text+=page.extract_text() or ""

        result=convert(text,mode)

        open("result.txt","w",encoding="utf-8").write(result)

        await message.answer_document(types.FSInputFile("result.txt"))


# ================= TEXT =================

@dp.message()
async def text_handler(message: types.Message):

    user_id=message.from_user.id

    mode=user_mode.get(user_id,"kl")

    result=convert(message.text,mode)

    await message.answer(result)


# ================= RUN =================

async def main():
    await dp.start_polling(bot)


if __name__=="__main__":
    asyncio.run(main())
